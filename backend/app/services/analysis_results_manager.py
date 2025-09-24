# backend/app/services/analysis_results_manager.py

import asyncio
import logging
import io
from typing import Dict, List, Any, Optional
from datetime import datetime, timezone
from dataclasses import dataclass, asdict

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook

from google.cloud import firestore
from google.cloud import storage
from google.cloud import pubsub_v1

from ..core.gcp_config import GCPSettings
from ..models.analysis_models import AnalysisResults, ExportRequest
from ..middleware.monitoring import performance_monitor

logger = logging.getLogger(__name__)

@dataclass
class AnalysisResultsSummary:
    """Summary of analysis results for dashboard display"""
    session_id: str
    topic: str
    user_id: str
    status: str
    overall_score: float
    confidence: float
    layer_scores: Dict[str, float]
    factor_scores: Dict[str, float]
    segment_scores: Dict[str, float]
    insights: List[str]
    recommendations: List[str]
    created_at: str
    completed_at: str
    processing_time: float

class AnalysisResultsManager:
    """Manage analysis results retrieval, formatting, and export"""
    
    def __init__(self):
        self.settings = GCPSettings()
        self.firestore_client = firestore.Client()
        self.storage_client = storage.Client()
        self.publisher = pubsub_v1.PublisherClient()
        
    @performance_monitor
    async def get_analysis_results(self, session_id: str) -> AnalysisResults:
        """Get complete analysis results for a session"""
        
        logger.info(f"Retrieving analysis results for session {session_id}")
        
        try:
            # Get session information
            session_ref = self.firestore_client.collection('analysis_sessions').document(session_id)
            session_doc = await session_ref.get()
            
            if not session_doc.exists:
                raise ValueError(f"Analysis session {session_id} not found")
            
            session_data = session_doc.to_dict()
            
            # Get layer scores
            layer_scores = await self._get_layer_scores(session_id)
            
            # Get factor calculations
            factor_calculations = await self._get_factor_calculations(session_id)
            
            # Get segment scores
            segment_scores = await self._get_segment_scores(session_id)
            
            # Generate comprehensive insights
            comprehensive_insights = await self._generate_comprehensive_insights(
                layer_scores, factor_calculations, segment_scores
            )
            
            # Calculate overall metrics
            overall_metrics = await self._calculate_overall_metrics(
                layer_scores, factor_calculations, segment_scores
            )
            
            results = AnalysisResults(
                session_id=session_id,
                topic=session_data.get('topic'),
                user_id=session_data.get('user_id'),
                status=session_data.get('status'),
                layer_scores=layer_scores,
                factor_calculations=factor_calculations,
                segment_scores=segment_scores,
                overall_metrics=overall_metrics,
                insights=comprehensive_insights['insights'],
                recommendations=comprehensive_insights['recommendations'],
                metadata={
                    'analysis_timestamp': session_data.get('completed_at'),
                    'processing_time': session_data.get('processing_time', 0),
                    'total_documents_analyzed': session_data.get('total_documents', 0),
                    'quality_threshold': session_data.get('quality_threshold', 0.6)
                }
            )
            
            logger.info(f"✅ Retrieved complete analysis results for session {session_id}")
            return results
            
        except Exception as e:
            logger.error(f"Failed to retrieve analysis results: {e}")
            raise

    async def get_results_summary(self, user_id: str, limit: int = 20) -> List[AnalysisResultsSummary]:
        """Get summary of analysis results for a user"""
        
        try:
            # Query user's analysis sessions
            sessions_ref = self.firestore_client.collection('analysis_sessions')
            query = sessions_ref.where('user_id', '==', user_id)\
                              .where('status', '==', 'completed')\
                              .order_by('completed_at', direction=firestore.Query.DESCENDING)\
                              .limit(limit)
            
            sessions = await query.get()
            
            summaries = []
            for session_doc in sessions:
                session_data = session_doc.to_dict()
                session_id = session_doc.id
                
                # Get quick metrics
                layer_scores = await self._get_layer_scores_summary(session_id)
                factor_scores = await self._get_factor_scores_summary(session_id)
                segment_scores = await self._get_segment_scores_summary(session_id)
                
                # Create summary
                summary = AnalysisResultsSummary(
                    session_id=session_id,
                    topic=session_data.get('topic'),
                    user_id=session_data.get('user_id'),
                    status=session_data.get('status'),
                    overall_score=session_data.get('overall_score', 0.0),
                    confidence=session_data.get('overall_confidence', 0.0),
                    layer_scores=layer_scores,
                    factor_scores=factor_scores,
                    segment_scores=segment_scores,
                    insights=session_data.get('key_insights', [])[:3],  # Top 3 insights
                    recommendations=session_data.get('key_recommendations', [])[:3],
                    created_at=session_data.get('created_at'),
                    completed_at=session_data.get('completed_at'),
                    processing_time=session_data.get('processing_time', 0)
                )
                
                summaries.append(summary)
            
            return summaries
            
        except Exception as e:
            logger.error(f"Failed to get results summary: {e}")
            return []

    async def export_results(self, session_id: str, format: str, user_id: str) -> Dict[str, Any]:
        """Export analysis results in specified format"""
        
        logger.info(f"Exporting results for session {session_id} in {format} format")
        
        try:
            # Get complete results
            results = await self.get_analysis_results(session_id)
            
            # Export based on format
            if format.lower() == 'pdf':
                export_result = await self._export_to_pdf(results, user_id)
            elif format.lower() == 'excel':
                export_result = await self._export_to_excel(results, user_id)
            elif format.lower() == 'json':
                export_result = await self._export_to_json(results, user_id)
            else:
                raise ValueError(f"Unsupported export format: {format}")
            
            # Log export event
            await self._log_export_event(session_id, format, user_id, export_result)
            
            return export_result
            
        except Exception as e:
            logger.error(f"Failed to export results: {e}")
            raise

    async def _export_to_pdf(self, results: AnalysisResults, user_id: str) -> Dict[str, Any]:
        """Export results to PDF format"""
        try:
            # For now, return a mock PDF export
            # In production, you would use reportlab or similar to generate actual PDF
            filename = f"exports/{user_id}/{results.session_id}/analysis_report.pdf"
            bucket = self.storage_client.bucket(f"{self.settings.project_id}-{self.settings.storage_bucket_prefix}")
            blob = bucket.blob(filename)
            
            # Create PDF using reportlab
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            
            # Title
            p.setFont("Helvetica-Bold", 16)
            p.drawString(100, height - 100, "Validatus Analysis Report")
            
            # Session info
            p.setFont("Helvetica", 12)
            p.drawString(100, height - 140, f"Session: {results.session_id}")
            p.drawString(100, height - 160, f"Topic: {results.topic}")
            p.drawString(100, height - 180, f"Generated: {datetime.now(timezone.utc).isoformat()}")
            
            # Overall metrics
            y_pos = height - 220
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_pos, "Overall Metrics")
            
            y_pos -= 30
            p.setFont("Helvetica", 12)
            overall_score = results.overall_metrics.get('overall_score', 0)
            confidence = results.overall_metrics.get('confidence', 0)
            p.drawString(120, y_pos, f"Overall Score: {overall_score:.2f}")
            y_pos -= 20
            p.drawString(120, y_pos, f"Confidence: {confidence:.2f}")
            
            # Layer scores
            y_pos -= 40
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_pos, "Layer Scores")
            
            y_pos -= 30
            p.setFont("Helvetica", 12)
            for score in results.layer_scores[:10]:  # Limit to first 10
                if y_pos < 100:  # Start new page if needed
                    p.showPage()
                    y_pos = height - 100
                p.drawString(120, y_pos, f"{score.layer_name}: {score.score:.2f}")
                y_pos -= 20
            
            # Key insights
            y_pos -= 40
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_pos, "Key Insights")
            
            y_pos -= 30
            p.setFont("Helvetica", 12)
            for insight in results.insights[:5]:
                if y_pos < 100:
                    p.showPage()
                    y_pos = height - 100
                p.drawString(120, y_pos, f"• {insight}")
                y_pos -= 20
            
            # Recommendations
            y_pos -= 40
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_pos, "Recommendations")
            
            y_pos -= 30
            p.setFont("Helvetica", 12)
            for rec in results.recommendations[:5]:
                if y_pos < 100:
                    p.showPage()
                    y_pos = height - 100
                p.drawString(120, y_pos, f"• {rec}")
                y_pos -= 20
            
            p.save()
            buffer.seek(0)
            
            # Upload to GCS
            blob.upload_from_file(buffer, content_type='application/pdf')
            
            size = buffer.getbuffer().nbytes
            
            return {
                'format': 'pdf',
                'filename': filename,
                'download_url': blob.public_url,
                'size': size,
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            raise

    async def _export_to_excel(self, results: AnalysisResults, user_id: str) -> Dict[str, Any]:
        """Export results to Excel format"""
        try:
            # For now, return a mock Excel export
            # In production, you would use openpyxl to create actual Excel files
            filename = f"exports/{user_id}/{results.session_id}/analysis_data.xlsx"
            bucket = self.storage_client.bucket(f"{self.settings.project_id}-{self.settings.storage_bucket_prefix}")
            blob = bucket.blob(filename)
            
            # Create Excel file using openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Results"
            
            # Add session metadata
            ws['A1'] = 'Session ID'
            ws['B1'] = results.session_id
            ws['A2'] = 'Topic'
            ws['B2'] = results.topic
            ws['A3'] = 'Generated'
            ws['B3'] = datetime.now(timezone.utc).isoformat()
            
            # Add layer scores
            ws['A5'] = 'Layer Scores'
            ws['A6'] = 'Layer Name'
            ws['B6'] = 'Score'
            ws['C6'] = 'Confidence'
            
            row = 7
            for score in results.layer_scores:
                ws[f'A{row}'] = score.layer_name
                ws[f'B{row}'] = score.score
                ws[f'C{row}'] = score.confidence
                row += 1
            
            # Add insights
            ws[f'A{row+1}'] = 'Key Insights'
            row += 2
            for insight in results.insights:
                ws[f'A{row}'] = insight
                row += 1
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Upload to GCS
            blob.upload_from_file(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            size = buffer.getbuffer().nbytes
            
            return {
                'format': 'excel',
                'filename': filename,
                'download_url': blob.public_url,
                'size': size,
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise

    async def _export_to_json(self, results: AnalysisResults, user_id: str) -> Dict[str, Any]:
        """Export results to JSON format"""
        try:
            # Convert results to JSON-serializable format
            json_data = {
                'session_id': results.session_id,
                'topic': results.topic,
                'user_id': results.user_id,
                'status': results.status,
                'analysis_results': {
                    'layer_scores': [asdict(score) for score in results.layer_scores],
                    'factor_calculations': [asdict(calc) for calc in results.factor_calculations],
                    'segment_scores': [asdict(score) for score in results.segment_scores],
                    'overall_metrics': results.overall_metrics,
                    'insights': results.insights,
                    'recommendations': results.recommendations
                },
                'metadata': results.metadata,
                'export_timestamp': datetime.now(timezone.utc).isoformat()
            }
            
            # Store in Cloud Storage
            filename = f"exports/{user_id}/{results.session_id}/analysis_results.json"
            bucket = self.storage_client.bucket(f"{self.settings.project_id}-{self.settings.storage_bucket_prefix}")
            blob = bucket.blob(filename)
            
            import json
            blob.upload_from_string(json.dumps(json_data, indent=2), content_type='application/json')
            
            return {
                'format': 'json',
                'filename': filename,
                'download_url': blob.public_url,
                'size': len(json.dumps(json_data)),
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            raise

    async def get_real_time_progress(self, session_id: str) -> Dict[str, Any]:
        """Get real-time progress updates for an analysis session"""
        
        try:
            # Get progress from Firestore
            progress_ref = self.firestore_client.collection('analysis_progress').document(session_id)
            progress_doc = await progress_ref.get()
            
            if not progress_doc.exists:
                return {'status': 'not_found'}
            
            progress_data = progress_doc.to_dict()
            
            return {
                'session_id': session_id,
                'status': progress_data.get('status'),
                'current_stage': progress_data.get('current_stage'),
                'progress_percentage': progress_data.get('progress_percentage', 0),
                'completed_layers': progress_data.get('completed_layers', []),
                'completed_factors': progress_data.get('completed_factors', []),
                'completed_segments': progress_data.get('completed_segments', []),
                'estimated_completion': progress_data.get('estimated_completion'),
                'last_updated': progress_data.get('last_updated'),
                'error_messages': progress_data.get('error_messages', [])
            }
            
        except Exception as e:
            logger.error(f"Failed to get real-time progress: {e}")
            return {'status': 'error', 'message': str(e)}

    async def _get_layer_scores(self, session_id: str) -> List[Any]:
        """Get layer scores for a session"""
        try:
            layer_scores_ref = self.firestore_client.collection('layer_scores')
            query = layer_scores_ref.where('session_id', '==', session_id)
            docs = await query.get()
            
            return [doc.to_dict() for doc in docs]
        except Exception as e:
            logger.error(f"Failed to get layer scores: {e}")
            return []

    async def _get_factor_calculations(self, session_id: str) -> List[Any]:
        """Get factor calculations for a session"""
        try:
            factor_calc_ref = self.firestore_client.collection('factor_calculations')
            query = factor_calc_ref.where('session_id', '==', session_id)
            docs = await query.get()
            
            return [doc.to_dict() for doc in docs]
        except Exception as e:
            logger.error(f"Failed to get factor calculations: {e}")
            return []

    async def _get_segment_scores(self, session_id: str) -> List[Any]:
        """Get segment scores for a session"""
        try:
            segment_scores_ref = self.firestore_client.collection('segment_scores')
            query = segment_scores_ref.where('session_id', '==', session_id)
            docs = await query.get()
            
            return [doc.to_dict() for doc in docs]
        except Exception as e:
            logger.error(f"Failed to get segment scores: {e}")
            return []

    async def _get_layer_scores_summary(self, session_id: str) -> Dict[str, float]:
        """Get layer scores summary"""
        layer_scores = await self._get_layer_scores(session_id)
        return {score.get('layer_name', ''): score.get('score', 0.0) for score in layer_scores}

    async def _get_factor_scores_summary(self, session_id: str) -> Dict[str, float]:
        """Get factor scores summary"""
        factor_calcs = await self._get_factor_calculations(session_id)
        return {calc.get('factor_name', ''): calc.get('score', 0.0) for calc in factor_calcs}

    async def _get_segment_scores_summary(self, session_id: str) -> Dict[str, float]:
        """Get segment scores summary"""
        segment_scores = await self._get_segment_scores(session_id)
        return {score.get('segment_name', ''): score.get('attractiveness_score', 0.0) for score in segment_scores}

    async def _generate_comprehensive_insights(self, layer_scores, factor_calculations, segment_scores):
        """Generate comprehensive insights from analysis results"""
        insights = []
        recommendations = []
        
        # Generate insights based on layer scores
        if layer_scores:
            avg_score = sum(score.get('score', 0) for score in layer_scores) / len(layer_scores)
            if avg_score > 0.8:
                insights.append("Strong performance across all strategic layers")
                recommendations.append("Maintain current strategic focus and continue optimization")
            elif avg_score < 0.5:
                insights.append("Significant opportunities for improvement across multiple layers")
                recommendations.append("Develop comprehensive improvement strategy")
        
        # Generate insights based on factor calculations
        if factor_calculations:
            high_confidence_factors = [calc for calc in factor_calculations if calc.get('confidence', 0) > 0.8]
            if high_confidence_factors:
                insights.append(f"{len(high_confidence_factors)} factors show high confidence levels")
                recommendations.append("Leverage high-confidence factors for strategic decisions")
        
        return {
            'insights': insights[:10],  # Limit to top 10 insights
            'recommendations': recommendations[:10]  # Limit to top 10 recommendations
        }

    async def _calculate_overall_metrics(self, layer_scores, factor_calculations, segment_scores):
        """Calculate overall metrics from analysis results"""
        overall_score = 0.0
        confidence = 0.0
        
        if layer_scores:
            overall_score = sum(score.get('score', 0) for score in layer_scores) / len(layer_scores)
            confidence = sum(score.get('confidence', 0) for score in layer_scores) / len(layer_scores)
        
        return {
            'overall_score': overall_score,
            'confidence': confidence,
            'total_layers': len(layer_scores),
            'total_factors': len(factor_calculations),
            'total_segments': len(segment_scores)
        }

    async def _log_export_event(self, session_id: str, format: str, user_id: str, export_result: Dict[str, Any]):
        """Log export event for audit trail"""
        try:
            export_log = {
                'session_id': session_id,
                'user_id': user_id,
                'format': format,
                'export_timestamp': datetime.now(timezone.utc).isoformat(),
                'filename': export_result.get('filename'),
                'size': export_result.get('size'),
                'download_url': export_result.get('download_url')
            }
            
            # Store in Firestore
            export_log_ref = self.firestore_client.collection('export_logs').document()
            await export_log_ref.set(export_log)
            
        except Exception as e:
            logger.error(f"Failed to log export event: {e}")

# Export the class
__all__ = ['AnalysisResultsManager', 'AnalysisResultsSummary']
